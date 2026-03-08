"""Excel workbook export for analyst delivery."""
from __future__ import annotations

import dataclasses
import logging
from pathlib import Path
from typing import Any

from src.pipeline.run_context import RunContext
from src.utils.file_utils import ensure_dir

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False
    logger.warning("openpyxl not available — Excel export disabled")


def _to_dict(obj: Any) -> dict:
    if dataclasses.is_dataclass(obj) and not isinstance(obj, type):
        d = dataclasses.asdict(obj)
        for k, v in d.items():
            if isinstance(v, list):
                d[k] = "; ".join(str(x) for x in v)
        return d
    return dict(obj)


def _write_sheet(wb, title: str, rows: list[Any]) -> None:
    """Add a sheet to *wb* with *rows* (dataclass instances)."""
    ws = wb.create_sheet(title=title[:31])  # Excel sheet names max 31 chars
    if not rows:
        ws.append(["(no data)"])
        return
    dicts = [_to_dict(r) for r in rows]
    headers = list(dicts[0].keys())
    ws.append(headers)
    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row_dict in dicts:
        ws.append([row_dict.get(h, "") for h in headers])


def export_excel(ctx: RunContext, filename: str = "guest_hub_delivery.xlsx") -> Path | None:
    """Write a multi-sheet Excel workbook to the hub output directory.

    Returns the output path, or None when openpyxl is not available.
    """
    if not _OPENPYXL_OK:
        logger.warning("Skipping Excel export — openpyxl not available")
        return None

    out_path = ensure_dir(ctx.hub_dir) / filename
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _write_sheet(wb, "rooms_canonical",         ctx.rooms_canonical)
    _write_sheet(wb, "spa_canonical",           ctx.spa_canonical)
    _write_sheet(wb, "dining_canonical",        ctx.dining_canonical)
    _write_sheet(wb, "dim_guest",               list(ctx.dim_guests.values()))
    _write_sheet(wb, "dim_phone",               list(ctx.dim_phones.values()))
    _write_sheet(wb, "fact_room_stay",          ctx.fact_room_stays)
    _write_sheet(wb, "bridge_guest_activity",   ctx.bridge_guest_activity)
    _write_sheet(wb, "bridge_guest_room_stay",  ctx.bridge_guest_room_stay)
    _write_sheet(wb, "qa_name_issues",          ctx.qa_name_issues)
    _write_sheet(wb, "qa_phone_issues",         ctx.qa_phone_issues)
    _write_sheet(wb, "qa_possible_matches",     ctx.qa_possible_matches)
    _write_sheet(wb, "qa_unmatched_spa",        ctx.qa_unmatched_spa)
    _write_sheet(wb, "qa_unmatched_dining",     ctx.qa_unmatched_dining)

    wb.save(out_path)
    logger.info("Excel workbook written: %s", out_path)
    return out_path
